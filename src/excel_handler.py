import openpyxl
from config import FIXED_PW, RESULT_FILENAME

def load_accounts(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    accounts = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=True)):
        if row[0]:
            accounts.append({
                "id": str(row[0]).strip(),
                "pw": FIXED_PW,
                "order": idx
            })
    print(f"Loaded {len(accounts)} accounts from {excel_path}")
    return accounts

def init_result_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Result"
    ws.append([
        "ID", "Plan Info", "Total Data", "Remaining Data",
        "Giftable Data", "Gift Count",
        "Mobile Lines", "TV Lines", "WiFi Lines", "PPS",
        "Message Info", "Attempts", "Status", 
        "Additional Services", "Discount Programs", "Option Products",
        "Bill Details",
        "Failure Reason"
    ])
    return wb, ws

def save_results(wb, results, filename=RESULT_FILENAME):
    ws = wb.active
    for r in sorted(results, key=lambda x: x["order"]):
        ws.append(r["row"])
    wb.save(filename)
    print(f"Saved to {filename}")
